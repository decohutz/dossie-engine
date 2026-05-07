"""
Synthetic XLSX builder that reproduces the structural shape of the
Projeto Regenera CIM (which is confidential and cannot be committed
as a fixture). The numbers are fictitious; what matters is that the
file exercises the same code paths the real Regenera workbook does:

* Six entities — five operating (Alpha/Beta/Gamma/Delta/Epsilon),
  one non-operating shared-services unit (CSC) that the parser
  flags via the overhead/CSC heuristic.
* Year columns: 1 historical (2024) + 6 projected (2025-2030),
  with NO "E" suffix on the projected years — exercising the
  parser's strip-then-cutoff path that broke valuation in v007.
* The "Receita Liquida" label is written WITHOUT the accent on
  ``í``, exercising the accent-insensitive matchers in the
  scenario builder (E3.3, B7) and the pptx exporter (E5, P1.a).
* CSC has zero revenue and negative EBITDA, exercising the
  E3.3 fix that filters non-operating entities out of the
  consolidated aggregate.

Used by the Regenera regression test fixtures in conftest.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl


YEARS = ("2024", "2025", "2026", "2027", "2028", "2029", "2030")


def _add_op_entity(wb, name: str, base_revenue: float,
                   growth: float, ebitda_pct: float) -> None:
    """Add a DRE sheet for an operating entity with simple growth model."""
    ws = wb.create_sheet(f"DRE {name}")
    ws.cell(row=4, column=4, value=f"DRE {name}")
    for i, year in enumerate(YEARS):
        ws.cell(row=4, column=5 + i, value=year)

    revs = []
    rev = base_revenue
    for _ in YEARS:
        revs.append(rev)
        rev *= 1 + growth

    rows = [
        ("Receita Bruta",   *[round(r * 1.18, 2) for r in revs]),
        ("Impostos",        *[round(-r * 0.18, 2) for r in revs]),
        # NB: "Receita Liquida" intentionally without accent — exercises
        # the accent-insensitive matcher fix.
        ("Receita Liquida", *[round(r, 2) for r in revs]),
        ("CMV",             *[round(-r * 0.40, 2) for r in revs]),
        ("Lucro Bruto",     *[round(r * 0.60, 2) for r in revs]),
        ("SG&A",            *[round(-r * (0.60 - ebitda_pct), 2) for r in revs]),
        ("EBITDA",          *[round(r * ebitda_pct, 2) for r in revs]),
    ]
    for r, row_data in enumerate(rows, start=5):
        ws.cell(row=r, column=4, value=row_data[0])
        for c, val in enumerate(row_data[1:]):
            ws.cell(row=r, column=5 + c, value=val)


def _add_csc(wb) -> None:
    """Add a CSC (Centro de Serviços Compartilhados) sheet — non-
    operating: zero revenue, negative EBITDA tracking shared overhead.

    The parser flags this entity as ``non_operating=True`` based on
    the sheet-name pattern, which excludes it from the consolidated
    DRE used in valuation (E3.3 / B4 fix).
    """
    ws = wb.create_sheet("DRE CSC")
    ws.cell(row=4, column=4, value="DRE CSC")
    for i, year in enumerate(YEARS):
        ws.cell(row=4, column=5 + i, value=year)

    cost = 5_000.0
    costs = []
    for _ in YEARS:
        costs.append(round(cost, 2))
        cost *= 1.10

    rows = [
        ("Receita Bruta",   *(0,) * 7),
        ("Receita Liquida", *(0,) * 7),
        ("CMV",             *[-c for c in costs]),
        ("EBITDA",          *[-c for c in costs]),
    ]
    for r, row_data in enumerate(rows, start=5):
        ws.cell(row=r, column=4, value=row_data[0])
        for c, val in enumerate(row_data[1:]):
            ws.cell(row=r, column=5 + c, value=val)


def build_regenera_synthetic_xlsx(path: Path) -> Path:
    """Create the XLSX fixture at ``path`` and return it."""
    wb = openpyxl.Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "Confidential — synthetic test fixture"
    cover["A2"] = "Last Update"
    # 2025 → parser deduces last_actual_year=2024, matching YEARS[0]
    cover["B2"] = datetime(2025, 12, 1)

    # Five operating entities with varied growth + margin profiles
    _add_op_entity(wb, "Alpha",   base_revenue=80_000, growth=0.05, ebitda_pct=0.30)
    _add_op_entity(wb, "Beta",    base_revenue=10_000, growth=0.40, ebitda_pct=0.35)
    _add_op_entity(wb, "Gamma",   base_revenue= 1_000, growth=0.50, ebitda_pct=0.30)
    _add_op_entity(wb, "Delta",   base_revenue= 2_000, growth=0.30, ebitda_pct=0.25)
    _add_op_entity(wb, "Epsilon", base_revenue=    500, growth=0.45, ebitda_pct=0.32)
    _add_csc(wb)

    wb.save(path)
    return path

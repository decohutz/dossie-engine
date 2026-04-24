"""
Frank CIM regression tests.

These lock in the current behavior of the pipeline for the Projeto Frank CIM.
They are *golden-file* tests: the expected output is whatever the pipeline
produced on the day the baseline was committed (see tests/fixtures/frank_baseline/).

What these tests cover (what the refactor touches):
    - valuation.json:               full deep-equal
    - markdown financial section:   byte-for-byte
    - xlsx financial + valuation:   cell-by-cell

What they do NOT cover:
    - Company/market/transaction markdown prose:   LLM-variable, not our layer
    - PPTX visual content:                         matplotlib renders non-deterministically
    - Page-level classifier output:                internal, no product impact

When you intentionally change behavior, regenerate the baseline fixture and
commit the new golden files alongside the code change.
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from ._helpers import (
    deep_diff,
    extract_financial_md_section,
    first_cell_diff,
    read_sheet,
)


# ────────────────────────────────────────────────────────────────────
# valuation.json
# ────────────────────────────────────────────────────────────────────

def test_valuation_json_matches_baseline(frank_pipeline_outputs, frank_baseline_dir):
    """Every number, label, and scenario in the valuation output must match
    the baseline within float_tol=0.01.

    Checks: scenario EVs, DCF terminals, FCF projections, IRR/MOIC, inputs
    (WACC, multiples, stake, entry, net_debt), summaries, what_needs_to_be_true.
    """
    actual = json.loads(frank_pipeline_outputs["json"].read_text(encoding="utf-8"))
    expected = json.loads((frank_baseline_dir / "valuation.json").read_text(encoding="utf-8"))

    diffs = deep_diff(actual, expected, float_tol=0.01)

    if diffs:
        # Show at most 20 diffs — pytest truncation is aggressive
        preview = "\n".join(f"  {d}" for d in diffs[:20])
        more = f"\n  ... and {len(diffs) - 20} more" if len(diffs) > 20 else ""
        pytest.fail(
            f"valuation.json has {len(diffs)} diff(s) vs baseline:\n{preview}{more}"
        )


# ────────────────────────────────────────────────────────────────────
# markdown — financial section only
# ────────────────────────────────────────────────────────────────────

def test_md_financial_section_matches_baseline(
    frank_pipeline_outputs, frank_baseline_dir
):
    """The '## 2. Informações financeiras' section must be byte-identical.

    This is where the refactor's surface area lands in the markdown output:
    DRE tables iterated from financials.entities. Other sections vary with
    LLM output and aren't checked here.
    """
    actual_md = frank_pipeline_outputs["md"].read_text(encoding="utf-8")
    expected_md = (frank_baseline_dir / "dossie.md").read_text(encoding="utf-8")

    actual_section = extract_financial_md_section(actual_md)
    expected_section = extract_financial_md_section(expected_md)

    if actual_section != expected_section:
        # Locate the first divergence for a useful failure message
        for i, (a, e) in enumerate(zip(actual_section, expected_section)):
            if a != e:
                ctx_start = max(0, i - 40)
                ctx_end = i + 40
                pytest.fail(
                    f"md financial section diverges at char {i}:\n"
                    f"  actual  : ...{actual_section[ctx_start:ctx_end]!r}\n"
                    f"  expected: ...{expected_section[ctx_start:ctx_end]!r}"
                )
        pytest.fail(
            f"md financial section length differs: "
            f"{len(actual_section)} vs {len(expected_section)}"
        )


# ────────────────────────────────────────────────────────────────────
# xlsx — financial statements + valuation sheets
# ────────────────────────────────────────────────────────────────────

FINANCIAL_SHEETS = [
    "DRE Franqueadora",
    "DRE Distribuidora",
    "DRE Lojas Próprias",
    "Balanço Franqueadora",
    "Balanço Distribuidora",
    "Balanço Lojas Próprias",
    "Valuation",
]


def test_xlsx_sheet_names_match_baseline(
    frank_pipeline_outputs, frank_baseline_dir
):
    """The workbook structure (sheet names and order) must match the baseline."""
    actual_wb = load_workbook(frank_pipeline_outputs["xlsx"])
    expected_wb = load_workbook(frank_baseline_dir / "dossie.xlsx")

    assert actual_wb.sheetnames == expected_wb.sheetnames, (
        f"sheet names differ:\n"
        f"  actual   = {actual_wb.sheetnames}\n"
        f"  expected = {expected_wb.sheetnames}"
    )


@pytest.mark.parametrize("sheet_name", FINANCIAL_SHEETS)
def test_xlsx_sheet_cells_match_baseline(
    frank_pipeline_outputs, frank_baseline_dir, sheet_name
):
    """Each financial/valuation sheet must be cell-by-cell identical.

    Parametrized so each sheet is its own test case — when something regresses
    you see which sheet broke, not a wall of text.
    """
    actual_wb = load_workbook(frank_pipeline_outputs["xlsx"])
    expected_wb = load_workbook(frank_baseline_dir / "dossie.xlsx")

    if sheet_name not in actual_wb.sheetnames:
        pytest.fail(f"sheet {sheet_name!r} missing from actual workbook")
    if sheet_name not in expected_wb.sheetnames:
        pytest.fail(f"sheet {sheet_name!r} missing from baseline workbook")

    actual = read_sheet(actual_wb, sheet_name)
    expected = read_sheet(expected_wb, sheet_name)

    diff = first_cell_diff(actual, expected)
    if diff is not None:
        pytest.fail(f"sheet {sheet_name!r} differs: {diff}")


# ────────────────────────────────────────────────────────────────────
# sanity — entity model generalization
# ────────────────────────────────────────────────────────────────────

def test_financial_chapter_discovers_three_entities(frank_pipeline_outputs):
    """The Frank CIM has 3 entities. This guards against a refactor that
    silently loses one (or accidentally adds a fourth).
    """
    val = json.loads(frank_pipeline_outputs["json"].read_text(encoding="utf-8"))
    base_models = val["scenarios"]["base"]["models"]
    entity_names = [m["entity_name"] for m in base_models]
    assert entity_names == ["Franqueadora", "Distribuidora", "Lojas Próprias"], (
        f"Expected Frank's 3 entities in discovery order; got {entity_names}"
    )

"""
Regenera-synthetic CIM regression tests.

These lock in the current behavior of the pipeline for a Regenera-shaped
input — multi-entity (six entities, including a non-operating CSC),
projection-flagged years (no "E" suffix), accent-free "Receita Liquida"
labels, and manual override flags supplying the sector multiples.

The fixture is a synthetic XLSX (see ``tests/_regenera_synthetic.py``)
that reproduces the structural shape of the real Projeto Regenera CIM
without using the confidential numbers. The baseline files in
``tests/fixtures/regenera_synthetic_baseline/`` capture whatever the
pipeline produced when E6 was committed.

What these tests cover (the parts E3.x and E4 fixed, plus E3.5/E5):

    - 6 entities discovered, with CSC flagged non-operating
    - Consolidated DRE excludes CSC (E3.3 / B4)
    - Projected years correctly identified (E3.3 / B3, E4)
    - Accent-free 'Receita Liquida' parses (E3.3 / B7)
    - DRE block terminator works (E3.4 / B5)
    - Manual overrides populate global_multiples_median (E3.5)
    - IRR/MOIC compute and diverge across scenarios (E4)
    - XLSX produces 11 sheets including all 6 DREs + Valuation
    - markdown financial section matches byte-for-byte

What they don't cover:

    - PPTX content (matplotlib chart rendering is non-deterministic)
    - Web enrichment results (DDG flaky, not load-bearing here since
      ``enrich=False`` in the fixture)
    - Company / market / transaction prose (LLM-variable; ``use_llm=False``
      means those sections come out empty/minimal in the synthetic run)

When you intentionally change behavior on the Regenera-shape branch —
e.g. tweak how the consolidated DRE is built, change the override
provenance text, or re-tune scenario factors — regenerate the baseline
by deleting the fixtures directory and running this test once; the
session fixture writes fresh artifacts that you then commit as the
new golden files.
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from ._helpers import (
    deep_diff,
    extract_financial_md_section,
    first_cell_diff,
    read_sheet,
)


# ────────────────────────────────────────────────────────────────────
# valuation.json — full deep-equal
# ────────────────────────────────────────────────────────────────────

def test_regenera_valuation_json_matches_baseline(
    regenera_synthetic_pipeline_outputs, regenera_baseline_dir,
):
    """Every number, label, and scenario in the valuation output must
    match the baseline within float_tol=0.01. Covers: scenario EVs
    across all 4 methods (DCF perpetuity / DCF exit / EV/EBITDA /
    EV/Revenue), IRR/MOIC, inputs (WACC, override multiples, stake,
    entry, net_debt), summaries, what_needs_to_be_true.

    A failure here means the valuation engine produced different
    numbers than the baseline — most likely a regression in
    ``build_consolidated`` / ``build_scenarios`` / ``run_dcf`` /
    ``run_multiples`` / ``run_irr``.
    """
    actual = json.loads(
        regenera_synthetic_pipeline_outputs["json"].read_text(encoding="utf-8")
    )
    expected = json.loads(
        (regenera_baseline_dir / "valuation.json").read_text(encoding="utf-8")
    )

    diffs = deep_diff(actual, expected, float_tol=0.01)
    if diffs:
        # First 10 diffs is plenty for diagnosis without overwhelming
        # the output.
        msg = "valuation.json drift:\n" + "\n".join(f"  {d}" for d in diffs[:10])
        if len(diffs) > 10:
            msg += f"\n  ... and {len(diffs) - 10} more diffs"
        pytest.fail(msg)


# ────────────────────────────────────────────────────────────────────
# Markdown — financial section byte-for-byte
# ────────────────────────────────────────────────────────────────────

def test_regenera_md_financial_section_matches_baseline(
    regenera_synthetic_pipeline_outputs, regenera_baseline_dir,
):
    """The financial section (chapter 2) of the markdown must match
    byte-for-byte. With six entities including a non-operating CSC,
    this exercises:

    * The order entities appear in
    * The exact numeric formatting (commas, signs, em-dashes)
    * The DRE block terminator (no 'Premissas' rows leaking in)
    * The accent-free label preservation
    """
    actual_md = regenera_synthetic_pipeline_outputs["md"].read_text(encoding="utf-8")
    expected_md = (regenera_baseline_dir / "dossie.md").read_text(encoding="utf-8")

    actual_section = extract_financial_md_section(actual_md)
    expected_section = extract_financial_md_section(expected_md)

    if actual_section != expected_section:
        # Find first differing line for friendly error
        a_lines = actual_section.splitlines()
        e_lines = expected_section.splitlines()
        for i, (a, e) in enumerate(zip(a_lines, e_lines)):
            if a != e:
                pytest.fail(
                    f"Financial markdown drift at line {i + 1}:\n"
                    f"  actual:   {a!r}\n  expected: {e!r}"
                )
        # If lengths differ
        pytest.fail(
            f"Financial markdown line count: {len(a_lines)} vs {len(e_lines)}"
        )


# ────────────────────────────────────────────────────────────────────
# XLSX — sheet names + cell-by-cell
# ────────────────────────────────────────────────────────────────────

def test_regenera_xlsx_sheet_names_match_baseline(
    regenera_synthetic_pipeline_outputs, regenera_baseline_dir,
):
    """Sheet names + order. With 5 operating entities + 1 non-operating
    CSC + Visão Geral / Mercado / Transação / Gaps / Valuation, expect
    11 sheets. Drifts here mean either a sheet was renamed, an entity
    DRE went missing, or a new sheet was added without baseline update.
    """
    actual_wb = load_workbook(regenera_synthetic_pipeline_outputs["xlsx"], data_only=True)
    expected_wb = load_workbook(regenera_baseline_dir / "dossie.xlsx", data_only=True)

    assert actual_wb.sheetnames == expected_wb.sheetnames, (
        f"Sheet names drifted:\n"
        f"  actual:   {actual_wb.sheetnames}\n"
        f"  expected: {expected_wb.sheetnames}"
    )


# Sheets we cell-compare. Visão Geral is excluded because some fields
# (founding year, etc.) come from the LLM extractor and may shift slightly
# between runs even with use_llm=False (currently they don't, but the
# stability isn't formally guaranteed; financial sheets ARE the contract).
REGENERA_FINANCIAL_SHEETS = (
    "DRE Alpha",
    "DRE Beta",
    "DRE Gamma",
    "DRE Delta",
    "DRE Epsilon",
    "DRE CSC",
    "Valuation",
)


@pytest.mark.parametrize("sheet_name", REGENERA_FINANCIAL_SHEETS)
def test_regenera_xlsx_sheet_cells_match_baseline(
    sheet_name, regenera_synthetic_pipeline_outputs, regenera_baseline_dir,
):
    """For each financial sheet, every cell value must equal the
    baseline. Float-tolerant within 0.01.

    A drift here pinpoints which entity / which year / which line
    item changed. Six DREs + Valuation = 7 separately-parameterized
    tests, each diagnosing exactly one branch.
    """
    actual_wb = load_workbook(regenera_synthetic_pipeline_outputs["xlsx"], data_only=True)
    expected_wb = load_workbook(regenera_baseline_dir / "dossie.xlsx", data_only=True)

    if sheet_name not in actual_wb.sheetnames:
        pytest.skip(f"Sheet {sheet_name!r} not in actual workbook")
    if sheet_name not in expected_wb.sheetnames:
        pytest.skip(f"Sheet {sheet_name!r} not in baseline workbook")

    actual = read_sheet(actual_wb, sheet_name)
    expected = read_sheet(expected_wb, sheet_name)

    diff = first_cell_diff(actual, expected)
    if diff:
        pytest.fail(f"Sheet {sheet_name!r} drift: {diff}")


# ────────────────────────────────────────────────────────────────────
# Sanity — entities discovered + non-op flag
# ────────────────────────────────────────────────────────────────────

def test_regenera_financial_chapter_discovers_six_entities(
    regenera_synthetic_pipeline_outputs,
):
    """Six entities total — five operating + one non-operating (CSC).
    The CSC pattern detection in the parser is what flips ``non_operating``,
    and the consolidated builder filters on it.
    """
    dossier = regenera_synthetic_pipeline_outputs["dossier"]
    entities = dossier.financials.entities

    assert len(entities) == 6, (
        f"expected 6 entities; got {len(entities)}: {[e.name for e in entities]}"
    )

    operating = [e.name for e in entities if not e.non_operating]
    non_operating = [e.name for e in entities if e.non_operating]

    assert len(operating) == 5, f"expected 5 operating entities; got {operating}"
    assert non_operating == ["CSC"], (
        f"expected only CSC to be non-operating; got {non_operating}"
    )


def test_regenera_consolidated_excludes_csc(regenera_synthetic_pipeline_outputs):
    """The consolidated DRE in each scenario must exclude CSC's
    contribution. This is the load-bearing E3.3/B4 behavior — without
    it, CSC's negative EBITDA inflates losses and depresses the
    consolidated numbers.

    Pinned check: terminal EBITDA in Base scenario is positive
    (operating margin would otherwise be negative once CSC's overhead
    swallows the operating EBITDA from the small entities).
    """
    from src.valuation.scenarios import build_scenarios
    dossier = regenera_synthetic_pipeline_outputs["dossier"]
    engine = build_scenarios(dossier, verbose=False)
    base_terminal = engine.base.consolidated[-1]

    assert base_terminal.ebitda > 0, (
        f"Base terminal EBITDA should be positive after CSC exclusion; "
        f"got {base_terminal.ebitda:,.2f}"
    )


def test_regenera_consolidated_years_are_marked_projected(
    regenera_synthetic_pipeline_outputs,
):
    """Years 2025-2030 in the consolidated DRE must be flagged
    is_projected=True. This is the E4/B8 fix — without it the IRR
    engine sees an empty projection list and falls back to the
    Newton initial guess (15%) with MOIC=0.

    A failure here would also tank the IRR test below, but pinning
    the upstream invariant lets us diagnose root cause faster.
    """
    from src.valuation.scenarios import build_scenarios
    dossier = regenera_synthetic_pipeline_outputs["dossier"]
    engine = build_scenarios(dossier, verbose=False)

    for scenario in [engine.pessimistic, engine.base, engine.optimistic]:
        proj_years = [y.year for y in scenario.consolidated if y.is_projected]
        assert proj_years == ["2025", "2026", "2027", "2028", "2029", "2030"], (
            f"Scenario {scenario.name} projected years drifted: {proj_years}"
        )


def test_regenera_irr_moic_diverge_with_overrides(
    regenera_synthetic_pipeline_outputs,
):
    """With ev_ebitda_override=11.0 supplied, IRR and MOIC must compute
    to non-zero values and diverge between scenarios. This is the
    end-to-end check that the E3.5 override path → E4 IRR fix works.
    """
    val = json.loads(
        regenera_synthetic_pipeline_outputs["json"].read_text(encoding="utf-8")
    )
    by_name = {s["scenario_name"]: s for s in val["summaries"]}

    pess_irr = by_name["Pessimista"]["irr"]
    otim_irr = by_name["Otimista"]["irr"]
    pess_moic = by_name["Pessimista"]["moic"]
    otim_moic = by_name["Otimista"]["moic"]

    # All four numbers must be non-zero
    assert pess_irr != 0 and otim_irr != 0, "IRRs should not be zero"
    assert pess_moic != 0 and otim_moic != 0, "MOICs should not be zero"

    # Optimistic > pessimistic (same fix as E4 covers, pinned again here)
    assert otim_irr > pess_irr, (
        f"Optimistic IRR not greater than pessimistic: "
        f"Pess={pess_irr:.3f}, Otim={otim_irr:.3f}"
    )
    assert otim_moic > pess_moic, (
        f"Optimistic MOIC not greater than pessimistic: "
        f"Pess={pess_moic:.3f}, Otim={otim_moic:.3f}"
    )


def test_regenera_global_multiples_populated_via_override(
    regenera_synthetic_pipeline_outputs,
):
    """E3.5 invariant: when ev_ebitda_override and ev_revenue_override are
    supplied to run_pipeline, ``dossier.market.global_multiples_median``
    must be filled with both values + the manual-override provenance
    string. This is the entry point for the multiples-comparables and
    DCF-exit valuation methods.
    """
    dossier = regenera_synthetic_pipeline_outputs["dossier"]
    field = dossier.market.global_multiples_median

    assert field.is_filled, "global_multiples_median should be filled by override"
    val = field.value
    assert isinstance(val, dict), f"value should be dict, got {type(val).__name__}"
    assert val["ev_ebitda_median"] == pytest.approx(11.0)
    assert val["ev_revenue_median"] == pytest.approx(1.8)
    assert "manual" in val.get("source_note", "").lower(), (
        f"source_note should mention manual override; got {val.get('source_note')!r}"
    )

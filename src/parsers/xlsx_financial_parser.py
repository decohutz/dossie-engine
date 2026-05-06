"""
XLSX financial-statement parser.

Reads a financial workbook (DREs and optionally balance sheets, one per
business unit / entity) and produces the same shape that the PDF financial
parser produces: a populated FinancialChapter ready for the orchestrator
to consume.

Key design points
-----------------
* The parser is **layout-driven**, not Frank-specific. It scans every sheet
  whose name matches a configurable DRE/Balance heading regex, then for
  each match it auto-detects:
    - the entity name (from the "DRE <X>" cell on the heading row),
    - the year columns (numeric or "2030E"-style strings on the heading row),
    - whether the sheet carries a duplicated "Análise Vertical" / vertical-
      analysis block to the right (those columns are skipped).

* Numbers are converted from BRL absolute → BRL k (thousands), matching
  the convention used by the PDF parser and the rest of the pipeline.

* Lines that look like ratios within a DRE block (e.g. "margem bruta % RL",
  "Crescimento YoY (%)", "Variação IPCA YoY (%)") are filtered out — the
  valuation engine recomputes ratios from absolute values, and keeping them
  in `FinancialStatement.lines` would contaminate downstream consumers.

* `is_projected` is decided per-year using a single `last_actual_year`
  value (default: detected from a "Last Update" / "Ano-base" cell on the
  Cover sheet, falling back to the workbook's earliest year). Years
  strictly greater than `last_actual_year` are flagged projected; the
  trailing "E" suffix that some advisors put on the last year is stripped
  from the year label so downstream code sees consistent "2024", "2025", …

* Entities flagged as non-operating (CSC, "Centro de Serviços", shared-
  services-style sheets) are returned with `non_operating=True` so the
  valuation engine can skip them while the dossier still surfaces them.

* Parse problems are returned as a list of `XlsxParseIssue` so the caller
  decides how to surface them (warnings, gaps, hard fail). The parser
  itself does not raise on recoverable issues — a sheet that fails to
  parse is dropped from the output and reported.
"""
from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable

from ..models.evidence import Evidence
from ..models.financials import (
    FinancialChapter,
    FinancialEntity,
    FinancialLine,
    FinancialStatement,
)


# ── Sheet-name patterns ──────────────────────────────────────────────────
# Match "DRE Foo", "DRE - Foo", "DRE—Foo", with optional trailing whitespace.
# We DO NOT match "DRE por BU >" or " DRE por BU >" (section headers in the
# Regenera workbook) — those are short and end with ">", which the regex
# rejects via the {2,} repetition + the explicit ">" filter.
_DRE_SHEET_RE = re.compile(
    r"^\s*DRE\s*[-–—:]?\s*(?P<entity>[A-Za-zÀ-ÿ0-9][A-Za-zÀ-ÿ0-9 &/.\-_+]{1,})\s*$",
    re.IGNORECASE,
)
_BALANCE_SHEET_RE = re.compile(
    r"^\s*(?:BP|Balan[cç]o(?:\s+Patrimonial)?)\s*[-–—:]?\s*"
    r"(?P<entity>[A-Za-zÀ-ÿ0-9][A-Za-zÀ-ÿ0-9 &/.\-_+]{1,})\s*$",
    re.IGNORECASE,
)

# Sheets to always skip even if their name superficially matches.
_SKIP_NAME_PATTERNS = (
    re.compile(r">"),                    # section dividers like "DRE por BU >"
    re.compile(r"^\s*detalhamento\b", re.IGNORECASE),
    re.compile(r"^\s*support\b", re.IGNORECASE),
    re.compile(r"^\s*macro\b", re.IGNORECASE),
    re.compile(r"^\s*cover\b", re.IGNORECASE),
)

# Entity-name tokens that imply non-operating (overhead / shared services).
_NON_OPERATING_TOKENS = (
    "csc",
    "shared services",
    "centro de servi",          # "Centro de Serviços Compartilhados"
    "corporate overhead",
    "overhead",
    "holding",                   # ambiguous but flag-and-warn
    "elimina",                   # eliminações
)

# Consolidated DRE: parsed but stored separately in FinancialChapter.
_CONSOLIDATED_TOKENS = (
    "consolidado",
    "consolidated",
    "grupo",
    "group",
    "total",
    "company",
)

# Heuristics for the heading row — the row that carries year headers.
# We scan up to this many rows from the top of the sheet looking for a row
# that has at least 3 cells parseable as years (int 2000-2099 or "2030E").
_HEADING_SCAN_DEPTH = 12

# Lines we filter out of DRE blocks — these are ratios/derived metrics that
# get computed from line items by the valuation engine, not raw line items.
# Match by normalized label (accent- and case-insensitive substring).
_RATIO_LABEL_TOKENS = (
    "crescimento yoy",
    "variacao ipca",
    "variacao igpm",
    "variacao inpc",
    "margem bruta",
    "margem ebitda",
    "margem liquida",
    "margem ebit",
    "margem contribuicao",
    "% rl",
    "% rb",
    "% receita",
    "% da receita",
    "% do ebt",
    "% da rl",
    "anal. vertical",
    "analise vertical",
    "aliquota",
    "taxa de crescimento",
    "taxa efetiva",
)

# Cells we never treat as line labels.
_NOISE_LABEL_TOKENS = (
    "ref",
    "suporte",
    "support",
    "x",                         # Regenera puts an "x" in col B as a check mark
    "check",
    "verificacao",               # checksum row
)

# Once any of these labels is seen alone in the label column, the parser
# stops reading rows for the current statement. Brazilian financial packs
# often append a "Premissas" / "Notas" / "Drivers" section beneath the
# DRE proper, listing inputs (number of employees, average salary, etc.)
# that shouldn't be confused with line items. The XLSX parser previously
# read past this divider and surfaced rows like "Salário Médio ($)" or
# "Comissão por novo contrato ($)" as DRE lines, which polluted the
# dossier output and (for ratio-shaped values) inflated derived metrics.
_BLOCK_TERMINATOR_TOKENS = (
    "premissas",
    "notas",
    "drivers",
    "assumptions",
    "memoria de calculo",
    "memo de calculo",
)


# ── Public types ─────────────────────────────────────────────────────────
@dataclass
class XlsxParseIssue:
    """A non-fatal problem encountered during parsing.

    Severity is informational: the caller chooses what to do (log, surface
    as a Gap, fail the run, etc.). The parser itself never raises for these.
    """
    severity: str          # "warn" | "info" | "error"
    sheet: str             # source sheet name (or "" if global)
    message: str

    def __str__(self) -> str:
        prefix = f"[{self.severity.upper()}]"
        loc = f" {self.sheet}:" if self.sheet else ""
        return f"{prefix}{loc} {self.message}"


@dataclass
class XlsxParseResult:
    """What the parser produces.

    `chapter` is a *partial* FinancialChapter: it has entities and the
    consolidated DRE populated, but not the derived metrics, capex, or
    dividend projections — those are filled in by the orchestrator's
    existing post-processing.

    `last_actual_year` is what the parser detected from the workbook;
    callers can override and re-flag projected years if they have a
    better signal (e.g. from the CIM cover letter).
    """
    chapter: FinancialChapter
    issues: list[XlsxParseIssue] = field(default_factory=list)
    last_actual_year: int | None = None
    source_file: str = ""


# ── Top-level entry point ────────────────────────────────────────────────
def parse_xlsx_financials(
    path: str | Path,
    *,
    last_actual_year: int | None = None,
) -> XlsxParseResult:
    """Parse a financial-pack XLSX into a partial FinancialChapter.

    Parameters
    ----------
    path : str | Path
        Path to the XLSX file.
    last_actual_year : int, optional
        If supplied, overrides the parser's auto-detection: years > this
        are flagged projected, years <= are flagged actual. If omitted,
        the parser tries to detect it from the Cover sheet, then falls
        back to "earliest year in the data".

    Returns
    -------
    XlsxParseResult
        See dataclass docstring.
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - environment guard
        raise RuntimeError(
            "openpyxl is required to parse XLSX financials. "
            "Install with: pip install openpyxl"
        ) from exc

    src = str(path)
    issues: list[XlsxParseIssue] = []
    chapter = FinancialChapter()

    wb = openpyxl.load_workbook(src, data_only=True, read_only=False)

    # Step 1: figure out the actual-year cutoff if the caller didn't tell us.
    detected_cutoff = _detect_last_actual_year(wb, issues)
    cutoff = last_actual_year if last_actual_year is not None else detected_cutoff

    # Step 2: walk every sheet and route to the right handler.
    for sheet_name in wb.sheetnames:
        if _is_skip_sheet(sheet_name):
            continue

        ws = wb[sheet_name]

        # Try DRE first, then balance sheet.
        dre_match = _DRE_SHEET_RE.match(sheet_name)
        bal_match = _BALANCE_SHEET_RE.match(sheet_name) if not dre_match else None

        if not dre_match and not bal_match:
            continue

        is_dre = dre_match is not None
        entity_raw = (dre_match or bal_match).group("entity").strip()
        stmt = _parse_statement_sheet(
            ws,
            entity_raw=entity_raw,
            statement_type="dre" if is_dre else "balance_sheet",
            source_file=src,
            cutoff=cutoff,
            issues=issues,
        )

        if stmt is None or not stmt.lines:
            issues.append(XlsxParseIssue(
                severity="warn",
                sheet=sheet_name,
                message=f"sheet matched {'DRE' if is_dre else 'BS'} pattern but produced no parseable lines; skipped",
            ))
            continue

        # Route consolidated separately.
        if _is_consolidated_entity(entity_raw):
            if is_dre:
                if chapter.dre_consolidated is not None:
                    issues.append(XlsxParseIssue(
                        severity="warn", sheet=sheet_name,
                        message=f"multiple consolidated DREs found; keeping first, dropping '{entity_raw}'",
                    ))
                else:
                    chapter.dre_consolidated = stmt
            else:
                # Consolidated balance sheet — Frank doesn't have a slot for this in
                # FinancialChapter today. Surface as an issue and keep going.
                issues.append(XlsxParseIssue(
                    severity="info", sheet=sheet_name,
                    message="consolidated balance sheet found but FinancialChapter has no slot for it; ignored",
                ))
            continue

        # Per-entity routing.
        if is_dre:
            chapter.upsert_dre(entity_raw, stmt)
        else:
            chapter.upsert_balance(entity_raw, stmt)

        # Tag non-operating after upsert (which may have created the entity).
        if _is_non_operating_entity(entity_raw):
            ent = chapter.get_entity(entity_raw)
            if ent is not None and not ent.non_operating:
                ent.non_operating = True
                issues.append(XlsxParseIssue(
                    severity="info", sheet=sheet_name,
                    message=f"entity '{entity_raw}' flagged non-operating (matched overhead/shared-services pattern)",
                ))

    if not chapter.entities and chapter.dre_consolidated is None:
        issues.append(XlsxParseIssue(
            severity="error", sheet="",
            message=f"no parseable DRE/BS sheets found in workbook (checked {len(wb.sheetnames)} sheets)",
        ))

    return XlsxParseResult(
        chapter=chapter,
        issues=issues,
        last_actual_year=cutoff,
        source_file=src,
    )


# ── Sheet-level parsing ──────────────────────────────────────────────────
def _parse_statement_sheet(
    ws,
    *,
    entity_raw: str,
    statement_type: str,
    source_file: str,
    cutoff: int | None,
    issues: list[XlsxParseIssue],
) -> FinancialStatement | None:
    """Parse one DRE/BS sheet into a FinancialStatement."""
    # 1. Locate the heading row (the one with year headers).
    heading_row, year_columns = _locate_heading_row(ws)
    if heading_row is None or not year_columns:
        issues.append(XlsxParseIssue(
            severity="warn", sheet=ws.title,
            message="could not locate heading row with year columns",
        ))
        return None

    years_raw = [yc[1] for yc in year_columns]   # list of "2024", "2025", …, "2030E"
    years_clean = [_strip_year_suffix(y) for y in years_raw]
    is_projected = {
        clean: _is_year_projected(clean, raw, cutoff)
        for clean, raw in zip(years_clean, years_raw)
    }

    # 2. Identify the label column. It's the column on the heading row that
    #    holds the entity name (e.g. "DRE Laces"). For Regenera that's col D.
    label_col = _find_label_column(ws, heading_row, entity_raw)
    if label_col is None:
        issues.append(XlsxParseIssue(
            severity="warn", sheet=ws.title,
            message=f"could not locate label column on row {heading_row}",
        ))
        return None

    # 3. Walk from heading_row+1 downward, building lines.
    parsed: list[FinancialLine] = []
    last_data_row = max(yc[0] for yc in year_columns)  # all year cols on same row
    # Scan through the end of the sheet — sheets are short.
    for row_idx in range(heading_row + 1, ws.max_row + 1):
        label_cell = ws.cell(row=row_idx, column=label_col).value
        if label_cell is None:
            continue

        label = str(label_cell).strip()
        if not label or _is_noise_label(label):
            continue

        # If we hit a section divider like "Premissas" / "Notas" / "Drivers"
        # alone in the label column, stop reading. Anything below is
        # assumption metadata (rates, headcounts, % drivers), not DRE.
        if _is_block_terminator(label):
            break

        if _is_ratio_label(label):
            continue

        # Pull the values. If ALL are None/empty, skip the row (heading-style row).
        values_dict: dict[str, float] = {}
        proj_dict: dict[str, bool] = {}
        for (yr_row, yr_label), clean in zip(year_columns, years_clean):
            # We assume the year header row is the same for all year cols; pull
            # values from the *current* row in those columns.
            cell_val = ws.cell(row=row_idx, column=_year_col_index(year_columns, yr_label)).value
            num = _coerce_number(cell_val)
            if num is None:
                continue
            values_dict[clean] = num / 1000.0   # BRL → BRL k
            proj_dict[clean] = is_projected[clean]

        if not values_dict:
            continue

        # Detect ratio rows that slipped through the label filter (all values
        # are between -1 and 1 in absolute terms — typical for percentages
        # stored as fractions).
        if _looks_like_ratio_row(values_dict):
            continue

        parsed.append(FinancialLine(
            label=label,
            values=values_dict,
            is_projected=proj_dict,
            unit="BRL k",
            evidence=Evidence(
                source_file=source_file,
                page=None,
                excerpt=f"{ws.title}!R{row_idx}: {label}",
                confidence=0.95,
                extraction_method="xlsx_parse",
            ),
        ))

    if not parsed:
        return None

    # Best-effort entity-name normalization for downstream consumers: keep the
    # raw form (matches the `chapter.upsert_dre` call site).
    return FinancialStatement(
        entity_name=entity_raw,
        statement_type=statement_type,
        lines=parsed,
        years=years_clean,
        evidence=Evidence(
            source_file=source_file,
            page=None,
            excerpt=f"Parsed {len(parsed)} lines from sheet '{ws.title}' for {entity_raw}",
            confidence=0.9,
            extraction_method="xlsx_parse",
        ),
    )


# ── Heading detection ────────────────────────────────────────────────────
def _locate_heading_row(ws) -> tuple[int | None, list[tuple[int, str]]]:
    """Find the row that holds year column headers.

    Returns
    -------
    (row_index, year_columns)
        `year_columns` is a list of (column_index, year_label_string) for
        every cell on that row that parses as a year. If the parser sees
        a duplicated year run on the same row (typical for "Análise
        Vertical" blocks), it keeps only the *first* contiguous run.
    """
    # Limit reach: sheets can have hundreds of columns of stale formatting,
    # but real headers are in the first 30-ish.
    max_col = min(ws.max_column or 1, 40)

    for row_idx in range(1, min(ws.max_row + 1, _HEADING_SCAN_DEPTH + 1)):
        candidates: list[tuple[int, str]] = []
        for col_idx in range(1, max_col + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            year_str = _coerce_year_label(cell_val)
            if year_str is not None:
                candidates.append((col_idx, year_str))

        if len(candidates) >= 3:
            # If we have a duplicated run (e.g. years repeat for vertical
            # analysis), keep only the first contiguous block.
            return row_idx, _first_contiguous_run(candidates)

    return None, []


def _first_contiguous_run(year_cols: list[tuple[int, str]]) -> list[tuple[int, str]]:
    """Keep only the first contiguous run of year columns.

    Regenera DRE sheets duplicate the years for an "Análise Vertical"
    block to the right, separated by a "Ref" column. The two blocks have
    a column gap, so we cut at the first gap > 1.
    """
    if not year_cols:
        return []
    run = [year_cols[0]]
    for prev, curr in zip(year_cols, year_cols[1:]):
        if curr[0] - prev[0] == 1:
            run.append(curr)
        else:
            break
    return run


def _coerce_year_label(val) -> str | None:
    """Normalize a cell value to a year-label string ('2024', '2030E') or None."""
    if val is None:
        return None
    if isinstance(val, int) and 2000 <= val <= 2099:
        return str(val)
    if isinstance(val, float) and val.is_integer() and 2000 <= int(val) <= 2099:
        return str(int(val))
    if isinstance(val, str):
        s = val.strip()
        m = re.match(r"^(20\d{2})E?$", s)
        if m:
            return s
    if isinstance(val, datetime):
        return str(val.year) if 2000 <= val.year <= 2099 else None
    return None


def _strip_year_suffix(year: str) -> str:
    """Strip the trailing 'E' (estimate) marker — keep just the year digits."""
    return year[:-1] if year.endswith("E") else year


def _is_year_projected(clean_year: str, raw_year: str, cutoff: int | None) -> bool:
    """Decide whether a year column is projected.

    Two signals: (1) the raw label has trailing "E" (advisor said so);
    (2) the year is strictly greater than `cutoff` (the workbook's
    last actual year).
    """
    if raw_year.endswith("E"):
        return True
    if cutoff is None:
        return False  # conservative: we don't know, assume actual
    try:
        return int(clean_year) > cutoff
    except ValueError:
        return False


def _find_label_column(ws, heading_row: int, entity_raw: str) -> int | None:
    """Find the column on `heading_row` whose value matches the entity title.

    Two-pass match: first we look for a cell containing the entity name
    (so "DRE Laces" matches when entity is "Laces"). If that fails, we
    accept any cell that starts with "DRE " or "BP " — handles the case
    where the sheet name is "DRE Consolidado" but the in-sheet label is
    "DRE Grupo" (Regenera) or vice versa.
    """
    target = _normalize_text(entity_raw)
    max_col = min(ws.max_column or 1, 12)

    # Pass 1: exact entity-name match.
    for col_idx in range(1, max_col + 1):
        v = ws.cell(row=heading_row, column=col_idx).value
        if v is None:
            continue
        cell_norm = _normalize_text(str(v))
        if target and target in cell_norm:
            return col_idx

    # Pass 2: any "DRE <something>" / "BP <something>" / "Balanço <something>"
    # label. Useful when the sheet name and in-sheet label disagree on the
    # entity (common for the consolidated tab).
    label_prefixes = ("dre ", "bp ", "balanco ", "balanço ")
    for col_idx in range(1, max_col + 1):
        v = ws.cell(row=heading_row, column=col_idx).value
        if v is None:
            continue
        cell_norm = _normalize_text(str(v))
        if any(cell_norm.startswith(p) for p in label_prefixes):
            return col_idx

    return None


def _year_col_index(year_columns: list[tuple[int, str]], year_label: str) -> int:
    """Look up the column index for a given year label in the heading."""
    for col, lbl in year_columns:
        if lbl == year_label:
            return col
    raise KeyError(year_label)   # programmer error; year_columns built locally


# ── Numeric coercion ─────────────────────────────────────────────────────
def _coerce_number(val) -> float | None:
    """Convert a cell value to float, handling Brazilian formatting and None.

    Returns None for blank cells and unparseable strings — those rows are
    treated as having no datum for that year.
    """
    if val is None:
        return None
    if isinstance(val, bool):  # Excel sometimes coerces; treat as no-data
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and (val != val):  # NaN
            return None
        return float(val)
    if isinstance(val, str):
        s = val.strip()
        if not s or s in ("--", "—", "-", "n.a.", "n/a", "N/A"):
            return None
        # Brazilian: "1.234.567,89" → 1234567.89; "(1.234)" → -1234
        is_neg = False
        if s.startswith("(") and s.endswith(")"):
            is_neg = True
            s = s[1:-1].strip()
        s = s.replace("R$", "").replace(" ", "").strip()
        # Decide which char is the decimal separator:
        # 1. If both `,` and `.` appear, the LAST one is the decimal — typical
        #    of "1.234,56" (pt-BR) or "1,234.56" (US).
        # 2. If only `,` appears, it's the decimal (pt-BR "0,42").
        # 3. If only `.` appears, it's ambiguous: "400.000" is pt-BR thousands
        #    (= 400000), but "0.42" is US decimal (= 0.42). We disambiguate
        #    by checking whether the segment AFTER the last dot is exactly 3
        #    digits (= thousands separator) or not (= decimal point).
        last_dot = s.rfind(".")
        last_com = s.rfind(",")
        if last_dot >= 0 and last_com >= 0:
            if last_com > last_dot:
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        elif last_com >= 0:
            # comma only — distinguish thousands vs decimal by tail length
            tail = s[last_com + 1:]
            if len(tail) == 3 and tail.isdigit():
                # treat all commas as thousands separators (US-style)
                s = s.replace(",", "")
            else:
                # decimal comma (pt-BR "0,42")
                s = s.replace(",", ".")
        elif last_dot >= 0:
            # dot only — distinguish thousands vs decimal by tail length
            tail = s[last_dot + 1:]
            if len(tail) == 3 and tail.isdigit():
                # treat all dots as thousands separators
                s = s.replace(".", "")
            # else: leave as-is, float() reads it as decimal
        try:
            n = float(s)
            return -n if is_neg else n
        except ValueError:
            return None
    return None


# ── Label classification ─────────────────────────────────────────────────
def _normalize_text(s: str) -> str:
    """Case- and accent-insensitive normalization for label matching."""
    nfkd = unicodedata.normalize("NFKD", s or "")
    stripped = "".join(c for c in nfkd if not unicodedata.combining(c))
    return stripped.strip().lower()


def _is_ratio_label(label: str) -> bool:
    norm = _normalize_text(label)
    return any(tok in norm for tok in _RATIO_LABEL_TOKENS)


def _is_noise_label(label: str) -> bool:
    norm = _normalize_text(label).strip()
    return norm in _NOISE_LABEL_TOKENS


def _is_block_terminator(label: str) -> bool:
    """Whether this label marks the end of the DRE block proper.

    Brazilian financial packs typically append a 'Premissas' or 'Drivers'
    section listing assumption inputs (% rates, headcounts, salaries)
    beneath the DRE proper. We stop reading at the first such marker —
    everything below is metadata, not line items.

    Match is exact-after-normalization to avoid false positives: a line
    item like 'Premissas — CMV' (with extra qualifier) shouldn't trigger;
    only a bare 'Premissas' / 'Notas' / 'Drivers' acting as a section
    header does.
    """
    norm = _normalize_text(label).strip()
    return norm in _BLOCK_TERMINATOR_TOKENS


def _looks_like_ratio_row(values: dict[str, float]) -> bool:
    """Detect a ratio row by value distribution.

    We treat a row as a ratio if every non-zero value falls strictly within
    [-1.0, 1.0] (typical for percentages stored as decimal fractions, e.g.
    0.18 for an 18% margin). This catches ratio rows whose label slipped
    past `_is_ratio_label` (e.g. exotic translations or typos).
    """
    nz = [v for v in values.values() if v is not None and v != 0.0]
    if not nz:
        return False
    return all(-1.0 < v < 1.0 for v in nz)


def _is_skip_sheet(name: str) -> bool:
    return any(p.search(name) for p in _SKIP_NAME_PATTERNS)


def _is_consolidated_entity(entity: str) -> bool:
    norm = _normalize_text(entity)
    return any(tok in norm for tok in _CONSOLIDATED_TOKENS)


def _is_non_operating_entity(entity: str) -> bool:
    norm = _normalize_text(entity)
    return any(tok in norm for tok in _NON_OPERATING_TOKENS)


# ── Last-actual-year detection ───────────────────────────────────────────
def _detect_last_actual_year(wb, issues: list[XlsxParseIssue]) -> int | None:
    """Try to detect the last actual (non-projected) year from Cover.

    Looks for a "Last Update <date>" cell on the Cover sheet. The year
    of that date is treated as the last actual year. If the workbook has
    no Cover or no parseable date, returns None and the caller falls
    back to "no projection flag" (everything stays as-is).
    """
    for sheet_name in wb.sheetnames:
        if _normalize_text(sheet_name) != "cover":
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            for cell in row:
                if cell is None:
                    continue
                if isinstance(cell, datetime):
                    return cell.year - 1   # the year being reported as "last update" is the closing year; treat prior year as last actual
                if isinstance(cell, str):
                    m = re.search(r"last\s+update[^0-9]*(\d{4})", cell, re.IGNORECASE)
                    if m:
                        return int(m.group(1)) - 1
        break  # found Cover, no need to keep looking

    issues.append(XlsxParseIssue(
        severity="info", sheet="",
        message="no Cover/Last Update found; cannot auto-detect last actual year (everything will be flagged actual unless --last-actual-year passed)",
    ))
    return None


__all__ = [
    "parse_xlsx_financials",
    "XlsxParseIssue",
    "XlsxParseResult",
]

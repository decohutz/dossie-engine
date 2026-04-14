"""
Excel exporter for dossier data.
Generates a multi-sheet .xlsx with financial statements, company overview, and market data.
"""
from __future__ import annotations
import re
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from ..models.dossier import Dossier


# ── Style constants ──────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=10)
DATA_FONT = Font(name="Calibri", size=10)
TOTAL_FONT = Font(name="Calibri", bold=True, size=10)
TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
PROJECTED_FONT = Font(name="Calibri", size=10, italic=True, color="4472C4")
NEGATIVE_FONT = Font(name="Calibri", size=10, color="C00000")
NEGATIVE_PROJECTED = Font(name="Calibri", size=10, italic=True, color="C00000")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="B4C6E7"),
)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
LABEL_FONT = Font(name="Calibri", bold=True, size=10, color="2F5496")


def _auto_width(ws, min_width=10, max_width=22):
    """Auto-adjust column widths."""
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def _write_header_row(ws, row: int, values: list[str]):
    """Write a styled header row."""
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _is_total_line(label: str) -> bool:
    """Check if a line label represents a total/subtotal."""
    markers = ["(=)", "Receita Líquida", "Lucro Bruto", "EBITDA", "EBIT",
               "EBT", "Lucro Líquido", "Patrimônio Líquido", "Ativo", "Passivo"]
    return any(m in label for m in markers)


def _is_margin_line(label: str) -> bool:
    """Check if a line is a percentage margin."""
    return "Margem" in label or "%" in label


def _clean_financial_label(label: str) -> str:
    """Clean up labels that merged two lines from the PDF parser.

    Fixes cases like:
    - "(+/-) Outras Receitas/Despesas -- -- -- 2 31 -- -- -- -- -- Operacionais (=) EBITDA"
    - "(+/-) Outras Receitas/Despesas Não -- -- -- 75 10 -- -- -- -- -- Operacionais (+/-) Resultado Financeiro"
    """
    # Remove embedded numeric values that leaked from data columns (-- and digits)
    label = re.sub(r'\s+--(\s+--)*\s*', ' ', label)
    label = re.sub(r'\s+\d+(\s+\d+)*\s+', ' ', label)

    # If label contains (=) or (+/-) and is too long, it merged two lines
    if len(label) > 50:
        # Try to extract the last meaningful accounting label
        # Pattern: "noise... (=) EBITDA" → "(=) EBITDA"
        match = re.search(r'(\([=+/-]+\)\s*\w[\w\s&]*?)$', label)
        if match:
            label = match.group(1).strip()

    # Remove stray "Operacionais" or "Não Operacionais" fragments
    label = re.sub(r'\bOpera\w*\s*', '', label)
    label = re.sub(r'\bNão\s*$', '', label)

    # Clean up multiple spaces
    label = re.sub(r'\s+', ' ', label).strip()

    return label


def _write_financial_sheet(ws, stmt, sheet_title: str):
    """Write a financial statement (DRE or Balance Sheet) to a worksheet."""
    if stmt is None or not stmt.lines:
        ws.cell(row=1, column=1, value="Sem dados disponíveis")
        return

    years = stmt.years or []
    if not years:
        ws.cell(row=1, column=1, value="Sem dados disponíveis")
        return

    # Title
    ws.cell(row=1, column=1, value=sheet_title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(years) + 1)

    # Entity and type
    ws.cell(row=2, column=1, value=f"{stmt.entity_name} — {stmt.statement_type.upper()}")
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=10, italic=True, color="666666")

    # Header row
    header_row = 4
    headers = ["Linha"] + years
    _write_header_row(ws, header_row, headers)

    # Data rows
    for i, line in enumerate(stmt.lines):
        row = header_row + 1 + i
        label = _clean_financial_label(line.label)

        # Label cell
        label_cell = ws.cell(row=row, column=1, value=label)
        is_total = _is_total_line(label)
        is_margin = _is_margin_line(label)

        if is_total:
            label_cell.font = TOTAL_FONT
            label_cell.fill = TOTAL_FILL
        else:
            label_cell.font = DATA_FONT

        label_cell.border = THIN_BORDER

        # Value cells
        for j, year in enumerate(years):
            val = line.values.get(year)
            is_projected = line.is_projected.get(year, False)

            cell = ws.cell(row=row, column=j + 2)

            if val is not None:
                if is_margin:
                    cell.value = val
                    cell.number_format = '0.0%'
                else:
                    cell.value = val
                    cell.number_format = '#,##0'

                # Font styling
                if is_total:
                    cell.font = TOTAL_FONT
                    cell.fill = TOTAL_FILL
                elif is_projected and val < 0:
                    cell.font = NEGATIVE_PROJECTED
                elif is_projected:
                    cell.font = PROJECTED_FONT
                elif val < 0:
                    cell.font = NEGATIVE_FONT
                else:
                    cell.font = DATA_FONT
            else:
                cell.value = "—"
                cell.font = Font(name="Calibri", size=10, color="AAAAAA")

            cell.alignment = Alignment(horizontal="right")
            cell.border = THIN_BORDER

    # Unit note
    note_row = header_row + len(stmt.lines) + 2
    ws.cell(row=note_row, column=1, value=f"Unidade: {stmt.lines[0].unit if stmt.lines else 'BRL k'}")
    ws.cell(row=note_row, column=1).font = Font(name="Calibri", size=9, italic=True, color="999999")

    projected_note = "Itálico azul = projetado"
    ws.cell(row=note_row + 1, column=1, value=projected_note)
    ws.cell(row=note_row + 1, column=1).font = Font(name="Calibri", size=9, italic=True, color="4472C4")

    _auto_width(ws, min_width=14, max_width=18)
    ws.column_dimensions["A"].width = 40


def _write_overview_sheet(ws, dossier: Dossier):
    """Write the company overview sheet."""
    ws.cell(row=1, column=1, value="Dossiê — Visão Geral").font = TITLE_FONT
    ws.merge_cells("A1:D1")

    p = dossier.company.profile
    meta = dossier.metadata

    fields = [
        ("Projeto", meta.project_name),
        ("Empresa", meta.target_company),
        ("Razão Social", p.legal_name.value),
        ("Sede", p.headquarters.value),
        ("Fundação", p.founding_year.value),
        ("Setor", p.sector.value),
        ("Modelo de Negócio", p.business_model.value),
        ("Público-Alvo", p.target_audience.value),
        ("Nº Lojas", p.number_of_stores.value),
        ("Nº Funcionários", p.number_of_employees.value),
        ("", ""),
        ("DESCRIÇÃO", ""),
        ("", p.description.value),
    ]

    row = 3
    for label, val in fields:
        if label:
            ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        if val is not None:
            val_cell = ws.cell(row=row, column=2, value=str(val) if val else "—")
            val_cell.font = DATA_FONT
            val_cell.alignment = Alignment(wrap_text=True)
        row += 1

    # Executives
    row += 1
    ws.cell(row=row, column=1, value="DIRETORIA").font = TITLE_FONT
    row += 1
    _write_header_row(ws, row, ["Nome", "Cargo", "Participação (%)", "Background"])
    row += 1

    for exec in dossier.company.executives:
        ws.cell(row=row, column=1, value=exec.name).font = DATA_FONT
        ws.cell(row=row, column=2, value=exec.role).font = DATA_FONT
        pct_cell = ws.cell(row=row, column=3, value=exec.ownership_pct)
        pct_cell.font = DATA_FONT
        pct_cell.number_format = '0.0%' if exec.ownership_pct and exec.ownership_pct < 1 else '0.0'
        bg_cell = ws.cell(row=row, column=4, value=exec.background or "—")
        bg_cell.font = DATA_FONT
        bg_cell.alignment = Alignment(wrap_text=True)
        row += 1

    # Timeline
    row += 1
    ws.cell(row=row, column=1, value="HISTÓRICO").font = TITLE_FONT
    row += 1
    _write_header_row(ws, row, ["Ano", "Evento"])
    row += 1

    for event in sorted(dossier.company.timeline, key=lambda e: e.year):
        ws.cell(row=row, column=1, value=event.year).font = DATA_FONT
        ws.cell(row=row, column=2, value=event.description).font = DATA_FONT
        row += 1

    # Products
    row += 1
    ws.cell(row=row, column=1, value="PRODUTOS").font = TITLE_FONT
    row += 1
    _write_header_row(ws, row, ["Produto", "Categoria", "% Receita", "Próprio"])
    row += 1

    for prod in dossier.company.products:
        ws.cell(row=row, column=1, value=prod.name).font = DATA_FONT
        ws.cell(row=row, column=2, value=prod.category or "—").font = DATA_FONT
        rev_cell = ws.cell(row=row, column=3, value=prod.revenue_share_pct)
        rev_cell.font = DATA_FONT
        ws.cell(row=row, column=4, value="Sim" if prod.is_proprietary else "Não").font = DATA_FONT
        row += 1

    _auto_width(ws, min_width=12, max_width=50)
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["D"].width = 50


def _write_market_sheet(ws, dossier: Dossier):
    """Write market data sheet."""
    ws.cell(row=1, column=1, value="Mercado").font = TITLE_FONT
    ws.merge_cells("A1:D1")

    row = 3

    # Market sizes
    ws.cell(row=row, column=1, value="TAMANHO DE MERCADO").font = LABEL_FONT
    row += 1
    _write_header_row(ws, row, ["Geografia", "Valor", "Unidade", "Ano", "CAGR"])
    row += 1

    for ms in dossier.market.market_sizes:
        ws.cell(row=row, column=1, value=ms.geography or "—").font = DATA_FONT
        val_cell = ws.cell(row=row, column=2, value=ms.value if ms.value is not None else 0)
        val_cell.font = DATA_FONT
        val_cell.number_format = '#,##0.0'
        ws.cell(row=row, column=3, value=ms.unit or "").font = DATA_FONT
        ws.cell(row=row, column=4, value=ms.year).font = DATA_FONT
        cagr_cell = ws.cell(row=row, column=5, value=ms.cagr)
        cagr_cell.font = DATA_FONT
        if ms.cagr:
            cagr_cell.number_format = '0.0%'
        row += 1

    # Fragmentation
    row += 1
    if dossier.market.market_fragmentation.is_filled:
        ws.cell(row=row, column=1, value="Fragmentação").font = LABEL_FONT
        ws.cell(row=row, column=2, value=str(dossier.market.market_fragmentation.value)).font = DATA_FONT
        row += 2

    # Competitors
    ws.cell(row=row, column=1, value="CONCORRENTES").font = LABEL_FONT
    row += 1
    _write_header_row(ws, row, ["Nome", "Lojas", "Receita", "Unidade"])
    row += 1

    for comp in dossier.market.competitors:
        ws.cell(row=row, column=1, value=comp.name).font = DATA_FONT
        ws.cell(row=row, column=2, value=comp.stores).font = DATA_FONT
        rev_cell = ws.cell(row=row, column=3, value=comp.revenue)
        rev_cell.font = DATA_FONT
        rev_cell.number_format = '#,##0'
        ws.cell(row=row, column=4, value=comp.revenue_unit or "").font = DATA_FONT
        row += 1

    # Multiples
    row += 1
    if dossier.market.global_multiples_median.is_filled:
        ws.cell(row=row, column=1, value="MÚLTIPLOS GLOBAIS (Mediana)").font = LABEL_FONT
        row += 1
        multiples = dossier.market.global_multiples_median.value
        if isinstance(multiples, dict):
            for key, val in multiples.items():
                label = key.replace("_", " ").replace("median", "").strip().upper()
                ws.cell(row=row, column=1, value=label).font = DATA_FONT
                val_cell = ws.cell(row=row, column=2, value=val)
                val_cell.font = DATA_FONT
                val_cell.number_format = '0.0x'
                row += 1

    # Precedent transactions
    row += 1
    if dossier.market.precedent_transactions:
        ws.cell(row=row, column=1, value="TRANSAÇÕES PRECEDENTES").font = LABEL_FONT
        row += 1
        _write_header_row(ws, row, ["Comprador", "Alvo", "Valor", "EV/EBITDA", "Stake"])
        row += 1
        for tx in dossier.market.precedent_transactions:
            ws.cell(row=row, column=1, value=tx.buyer or "—").font = DATA_FONT
            ws.cell(row=row, column=2, value=tx.target or "—").font = DATA_FONT
            ws.cell(row=row, column=3, value=tx.value or "—").font = DATA_FONT
            val_cell = ws.cell(row=row, column=4, value=tx.ev_ebitda)
            val_cell.font = DATA_FONT
            if tx.ev_ebitda:
                val_cell.number_format = '0.0x'
            ws.cell(row=row, column=5, value=f"{tx.stake_pct}%" if tx.stake_pct else "—").font = DATA_FONT
            row += 1

    _auto_width(ws)


def _write_transaction_sheet(ws, dossier: Dossier):
    """Write transaction info sheet."""
    ws.cell(row=1, column=1, value="Transação").font = TITLE_FONT
    ws.merge_cells("A1:D1")

    t = dossier.transaction
    fields = [
        ("Contexto", t.context.value),
        ("Tipo", t.transaction_type.value),
        ("Stake Alvo", t.target_stake_range.value),
        ("Capital Necessário", t.capital_needed.value),
        ("OPEX", t.opex_component.value),
        ("CAPEX", t.capex_component.value),
        ("Uso dos Recursos", t.use_of_proceeds.value),
        ("Advisor", t.advisor.value),
        ("Perímetro", t.perimeter.value),
    ]

    row = 3
    for label, val in fields:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        val_cell = ws.cell(row=row, column=2, value=str(val) if val else "—")
        val_cell.font = DATA_FONT
        val_cell.alignment = Alignment(wrap_text=True)
        row += 1

    _auto_width(ws)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60


def _write_gaps_sheet(ws, dossier: Dossier):
    """Write gap analysis sheet."""
    ws.cell(row=1, column=1, value="Gap Analysis").font = TITLE_FONT
    ws.merge_cells("A1:E1")

    row = 3
    _write_header_row(ws, row, ["Severidade", "Capítulo", "Campo", "Descrição", "Fonte Sugerida", "Web?"])
    row += 1

    crit_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    imp_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for gap in dossier.gaps:
        sev_cell = ws.cell(row=row, column=1, value=gap.severity.upper())
        sev_cell.font = Font(name="Calibri", bold=True, size=10)
        sev_cell.fill = crit_fill if gap.severity == "critical" else imp_fill

        ws.cell(row=row, column=2, value=gap.chapter).font = DATA_FONT
        ws.cell(row=row, column=3, value=gap.field_path).font = DATA_FONT
        ws.cell(row=row, column=4, value=gap.description).font = DATA_FONT
        ws.cell(row=row, column=5, value=gap.suggested_source or "—").font = DATA_FONT
        ws.cell(row=row, column=6, value="🌐" if gap.requires_internet else "").font = DATA_FONT
        row += 1

    # Summary
    row += 1
    total = len(dossier.gaps)
    critical = sum(1 for g in dossier.gaps if g.severity == "critical")
    ws.cell(row=row, column=1, value=f"Total: {total} gaps ({critical} críticos)").font = LABEL_FONT

    _auto_width(ws)
    ws.column_dimensions["D"].width = 45


def export_xlsx(dossier: Dossier, output_path: str, verbose: bool = False) -> str:
    """Export dossier to a formatted Excel file.

    Args:
        dossier: The dossier to export
        output_path: Path for the output .xlsx file
        verbose: Print progress

    Returns:
        The output file path
    """
    if verbose:
        print("  [Excel] Gerando planilha...")

    wb = Workbook()

    # Sheet 1: Overview
    ws_overview = wb.active
    ws_overview.title = "Visão Geral"
    _write_overview_sheet(ws_overview, dossier)

    # Sheet 2-4: DREs
    fin = dossier.financials
    dre_sheets = [
        ("DRE Franqueadora", fin.dre_franqueadora),
        ("DRE Distribuidora", fin.dre_distribuidora),
        ("DRE Lojas Próprias", fin.dre_lojas_proprias),
    ]
    for title, stmt in dre_sheets:
        ws = wb.create_sheet(title=title)
        _write_financial_sheet(ws, stmt, title)

    # Sheet 5-7: Balance Sheets
    bal_sheets = [
        ("Balanço Franqueadora", fin.balance_franqueadora),
        ("Balanço Distribuidora", fin.balance_distribuidora),
        ("Balanço Lojas Próprias", fin.balance_lojas_proprias),
    ]
    for title, stmt in bal_sheets:
        ws = wb.create_sheet(title=title)
        _write_financial_sheet(ws, stmt, title)

    # Sheet 8: Market
    ws_market = wb.create_sheet(title="Mercado")
    _write_market_sheet(ws_market, dossier)

    # Sheet 9: Transaction
    ws_transaction = wb.create_sheet(title="Transação")
    _write_transaction_sheet(ws_transaction, dossier)

    # Sheet 10: Gaps
    ws_gaps = wb.create_sheet(title="Gaps")
    _write_gaps_sheet(ws_gaps, dossier)

    # Save
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    if verbose:
        print(f"  [Excel] ✅ Salvo em: {output_path}")

    return output_path